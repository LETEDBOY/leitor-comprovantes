"""
Leitor de Comprovantes -> Planilha de Prestação de Contas
-----------------------------------------------------------
Sobe fotos de comprovante (posto, lavanderia, etc.), a Claude API lê cada
imagem e extrai NF, DATA, DESCRIÇÃO e VALOR. Você revisa numa tabela editável
e exporta pro Excel no MESMO layout do seu template (linha 10 em diante,
coluna A=NF, B=DATA, C:G=DESCRIÇÃO mesclado, H=VALOR).

Como rodar:
    pip install streamlit anthropic openpyxl pandas --break-system-packages
    export ANTHROPIC_API_KEY=sua_chave_aqui
    streamlit run leitor_comprovantes.py
"""

import base64
import json
import os
from copy import copy
from datetime import date, datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from anthropic import Anthropic
import openpyxl

MODEL = "claude-sonnet-5"  # troque para "claude-haiku-4-5-20251001" se quiser mais barato/rápido

TEMPLATE_ROW_START = 10   # primeira linha de dados na sua planilha
TEMPLATE_ROW_END = 45     # última linha antes do TOTAL (=SUM(H10:H45) no template)

PROMPT = """Você está lendo uma foto de comprovante/cupom fiscal brasileiro
(pode ser posto de gasolina, lavanderia, ou outro estabelecimento).

Extraia e devolva APENAS um JSON válido, sem nenhum texto antes ou depois, com estes campos:
{
  "nf": "número da NF-e/NFC-e/comprovante (o número mais parecido com nota fiscal, ex: NFC-e nº, ou o código/ID do comprovante se não houver NF)",
  "data": "AAAA-MM-DD",
  "estabelecimento": "nome do estabelecimento",
  "descricao": "descrição curta do item/serviço, ex: 'Gasolina Comum (20,760 L)' ou 'Lavagem' ou 'Secagem'",
  "valor": 0.00
}

Regras:
- valor é o VALOR TOTAL PAGO, número com ponto decimal (não use vírgula).
- data no formato ISO AAAA-MM-DD.
- Se não conseguir ler algum campo com certeza, retorne null nesse campo, não invente.
- Não escreva nada fora do JSON.
"""


def extract_receipt(client: Anthropic, file_bytes: bytes, media_type: str) -> dict:
    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
    block_type = "document" if media_type == "application/pdf" else "image"
    resp = client.messages.create(
        model=MODEL,
        max_tokens=500,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": block_type, "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": PROMPT},
                ],
            }
        ],
    )
    text = "".join(block.text for block in resp.content if block.type == "text").strip()
    text = text.replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return {"nf": None, "data": None, "estabelecimento": None, "descricao": f"ERRO AO LER: {text[:200]}", "valor": None}


def build_description(row) -> str:
    est = (row.get("estabelecimento") or "").strip()
    desc = (row.get("descricao") or "").strip()
    if est and desc:
        return f"{est} - {desc}"
    return est or desc or ""


def fill_template(template_bytes: bytes, df: pd.DataFrame) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    ws = wb["Plan1"]

    # acha a primeira linha vazia no bloco de despesas (respeita o que já
    # estiver preenchido manualmente no template)
    row = TEMPLATE_ROW_START
    while row <= TEMPLATE_ROW_END and ws[f"A{row}"].value not in (None, ""):
        row += 1

    style_row = TEMPLATE_ROW_START  # usa a formatação da primeira linha como referência

    for _, r in df.iterrows():
        if row > TEMPLATE_ROW_END:
            st.warning(f"Planilha só tem espaço até a linha {TEMPLATE_ROW_END}. "
                       f"Alguns comprovantes ficaram de fora — copie manualmente ou aumente o template.")
            break
        for col in ["A", "B", "C", "H"]:
            src = ws[f"{col}{style_row}"]
            dst = ws[f"{col}{row}"]
            dst.font = copy(src.font)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
            dst.border = copy(src.border)

        ws[f"A{row}"] = r["NF"]
        if pd.notna(r["Data"]) and r["Data"]:
            try:
                ws[f"B{row}"] = datetime.strptime(str(r["Data"]), "%Y-%m-%d").date()
            except ValueError:
                ws[f"B{row}"] = r["Data"]
        ws[f"C{row}"] = r["Descrição"]
        ws[f"H{row}"] = float(r["Valor"]) if r["Valor"] not in (None, "") else None
        row += 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


st.set_page_config(page_title="Leitor de Comprovantes", page_icon="🧾", layout="wide")
st.title("🧾 Leitor de Comprovantes → Planilha")

api_key = os.environ.get("ANTHROPIC_API_KEY") or st.sidebar.text_input("Anthropic API key", type="password")
if not api_key:
    st.info("Cole sua ANTHROPIC_API_KEY na barra lateral (ou defina a variável de ambiente) pra começar.")
    st.stop()

client = Anthropic(api_key=api_key)

st.subheader("1. Suba as fotos dos comprovantes")
files = st.file_uploader(
    "Fotos ou PDFs (posto, lavanderia, etc.)",
    type=["jpg", "jpeg", "png", "pdf"],
    accept_multiple_files=True,
    help="No iPhone, ao tocar aqui o Safari oferece 'Tirar Foto', 'Escolher Foto' ou 'Escolher Arquivo' (pra PDF).",
)

if files and st.button(f"🔎 Ler {len(files)} comprovante(s)", type="primary"):
    results = []
    progress = st.progress(0.0, text="Lendo comprovantes...")
    for i, f in enumerate(files):
        if f.type == "application/pdf":
            media_type = "application/pdf"
        elif f.type == "image/png":
            media_type = "image/png"
        else:
            media_type = "image/jpeg"
        data = extract_receipt(client, f.getvalue(), media_type)
        results.append(
            {
                "Arquivo": f.name,
                "NF": data.get("nf") or "",
                "Data": data.get("data") or "",
                "Descrição": build_description(data),
                "Valor": data.get("valor"),
            }
        )
        progress.progress((i + 1) / len(files), text=f"Lendo {f.name}...")
    st.session_state["df"] = pd.DataFrame(results)
    progress.empty()

if "df" in st.session_state:
    st.subheader("2. Revise e corrija o que precisar")
    edited = st.data_editor(
        st.session_state["df"],
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Valor": st.column_config.NumberColumn(format="R$ %.2f"),
            "Data": st.column_config.TextColumn(help="AAAA-MM-DD"),
        },
        key="editor",
    )
    total = pd.to_numeric(edited["Valor"], errors="coerce").sum()
    st.metric("Total dos comprovantes revisados", f"R$ {total:,.2f}")

    st.subheader("3. Preencher a planilha")
    template_file = st.file_uploader("Sua planilha modelo (.xlsx)", type=["xlsx"], key="template")
    if template_file and st.button("📥 Gerar planilha preenchida", type="primary"):
        result_bytes = fill_template(template_file.getvalue(), edited)
        st.download_button(
            "Baixar planilha preenchida",
            data=result_bytes,
            file_name=f"prestacao_contas_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
